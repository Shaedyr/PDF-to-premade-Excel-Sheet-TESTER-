from openpyxl import load_workbook
from openpyxl.styles import Alignment
from io import BytesIO
import re

TARGET_FILL_HEX = "F2F2F2"   # Light gray fill used to mark fillable cells


# ---------------------------------------------------------
# Helper: Extract RGB hex from cell fill
# ---------------------------------------------------------
def _rgb_hex_from_color(col):
    if not col:
        return None

    rgb = getattr(col, "rgb", None)

    # Ignore theme colors, indexed colors, or missing RGB
    if not rgb or not isinstance(rgb, str):
        return None

    rgb = rgb.upper()

    # Ignore "AUTO" or other non-hex values
    if not all(c in "0123456789ABCDEF" for c in rgb if c != "A"):
        return None

    # Remove alpha channel (ARGB → RGB)
    if len(rgb) == 8:
        rgb = rgb[2:]

    return rgb if len(rgb) == 6 else None


# ---------------------------------------------------------
# Helper: Normalize label text
# ---------------------------------------------------------
def _normalize_label(t):
    return re.sub(r"[^a-zA-Z0-9æøåÆØÅ]+", " ", (t or "").lower()).strip()


# ---------------------------------------------------------
# FIELD KEYWORDS (mapping logic)
# ---------------------------------------------------------
FIELD_KEYWORDS = {
    "company_name": ["selskapsnavn", "navn", "firma"],
    "org_number": ["organisasjonsnummer", "orgnr", "org nr"],
    "address": ["adresse", "gate"],
    "post_nr": ["postnummer", "postnr"],
    "city": ["poststed", "by"],
    "employees": ["ansatte", "antall ansatte"],
    "homepage": ["hjemmeside", "nettside"],
    "nace_code": ["nacekode", "nace"],
    "nace_description": ["bransje", "næring"],
    "company_summary": ["om oss", "sammendrag"],
    "revenue_2024": ["omsetning"],
    "tender_deadline": ["frist", "anbudsfrist"],
}


# ---------------------------------------------------------
# STEP A: Scan template and find fillable cells
# ---------------------------------------------------------
def scan_template(template_bytes):
    wb = load_workbook(BytesIO(template_bytes), data_only=False)
    mapping = {}

    for ws in wb.worksheets:
        sheet_map = {}

        for row in ws.iter_rows():
            for cell in row:
                fg = getattr(cell.fill, "fgColor", None)
                hexcol = _rgb_hex_from_color(fg)

                if hexcol == TARGET_FILL_HEX:
                    label = None

                    # Try left cell
                    if cell.column > 1:
                        left = ws.cell(row=cell.row, column=cell.column - 1).value
                        if left:
                            label = str(left)

                    # Try above cell
                    if not label and cell.row > 1:
                        above = ws.cell(row=cell.row - 1, column=cell.column).value
                        if above:
                            label = str(above)

                    # Match label to field
                    if label:
                        norm = _normalize_label(label)
                        for field, kws in FIELD_KEYWORDS.items():
                            if any(kw in norm for kw in kws):
                                sheet_map[field] = cell.coordinate

        mapping[ws.title] = sheet_map

    return mapping


# ---------------------------------------------------------
# STEP B: Fill the workbook
# ---------------------------------------------------------
def fill_excel(template_bytes, field_values, summary_text):
    wb = load_workbook(BytesIO(template_bytes))
    mapping = scan_template(template_bytes)

    first_sheet = wb.sheetnames[0]
    ws_first = wb[first_sheet]

    # Fill all sheets
    for sheet_name, sheet_map in mapping.items():
        ws = wb[sheet_name]

        for field, coord in sheet_map.items():
            if field in field_values and field_values[field]:
                ws[coord].value = str(field_values[field])

    # Insert summary into first sheet
    if summary_text:
        placed = False

        for row in ws_first.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "skriv her" in cell.value.lower():
                    cell.value = summary_text
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    placed = True
                    break
            if placed:
                break

        # Fallback location
        if not placed:
            ws_first["A46"] = summary_text
            ws_first["A46"].alignment = Alignment(wrap_text=True, vertical="top")

    # Return final Excel bytes
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ---------------------------------------------------------
# PAGE VIEW (for debugging)
# ---------------------------------------------------------
def run():
    import streamlit as st
    st.title("📊 Excel Filler Module")
    st.write("Dette er et backend-modul som fyller Excel-malen.")
    st.info("Brukes av hovedsiden for å generere ferdig Excel.")
