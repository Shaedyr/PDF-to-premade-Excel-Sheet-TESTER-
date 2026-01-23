from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from io import BytesIO
from app_modules.Sheets.sheet_config import SHEET_MAPPINGS, transform_for_sheet
import streamlit as st

HEADLINE_COLORS = ["FF0BD7B5", "0BD7B5"]


def fill_excel(template_bytes, field_values, summary_text):
    """
    Fill Excel template with data from field_values.
    
    Args:
        template_bytes: Excel template file as bytes
        field_values: Dictionary of field values to fill
        summary_text: Company summary text
        
    Returns:
        Filled Excel file as bytes
    """
    wb = load_workbook(filename=BytesIO(template_bytes))

    # DEBUG: Show what data we received
    st.write("🔍 **EXCEL_FILLER DEBUG:**")
    st.write(f"📦 field_values keys: {list(field_values.keys())}")
    if "pdf_text" in field_values:
        st.success(f"✅ pdf_text EXISTS in field_values! Length: {len(field_values['pdf_text'])} chars")
    else:
        st.error("❌ pdf_text NOT in field_values!")
        st.write("Available keys:", list(field_values.keys()))

    # Process each sheet that has a mapping configured
    for sheet_name in SHEET_MAPPINGS.keys():
        if sheet_name not in wb.sheetnames:
            st.warning(f"⚠️ Sheet '{sheet_name}' not found in Excel template")
            continue

        ws = wb[sheet_name]
        
        st.write(f"📄 Processing sheet: {sheet_name}")
        
        # Transform data for this specific sheet
        transformed_data = transform_for_sheet(sheet_name, field_values)
        
        st.write(f"  ✓ Transformed data has {len(transformed_data)} fields")

        # Check if this sheet uses static mapping or dynamic mapping
        cell_map = SHEET_MAPPINGS.get(sheet_name, {})
        
        if cell_map:
            # STATIC MAPPING (like Sammendrag)
            # Uses predefined field_name -> cell_ref mapping
            st.write(f"  📌 Using static mapping ({len(cell_map)} mappings)")
            
            for field_key, cell_ref in cell_map.items():
                value = transformed_data.get(field_key, "")

                cell = ws[cell_ref]

                # Skip cells with headline colors (headers)
                fill = cell.fill
                if (
                    fill and isinstance(fill, PatternFill)
                    and fill.fgColor and fill.fgColor.rgb
                    and fill.fgColor.rgb.upper() in HEADLINE_COLORS
                ):
                    continue

                cell.value = value
                
        else:
            # DYNAMIC MAPPING (like Fordon)
            # transform_data() returns cell_ref -> value directly
            st.write(f"  📌 Using dynamic mapping ({len(transformed_data)} cells)")
            
            for cell_ref, value in transformed_data.items():
                # cell_ref is like "B3", "C3", etc.
                if isinstance(cell_ref, str) and len(cell_ref) >= 2:
                    try:
                        cell = ws[cell_ref]
                        
                        # Skip cells with headline colors (headers)
                        fill = cell.fill
                        if (
                            fill and isinstance(fill, PatternFill)
                            and fill.fgColor and fill.fgColor.rgb
                            and fill.fgColor.rgb.upper() in HEADLINE_COLORS
                        ):
                            continue
                        
                        cell.value = value
                        st.write(f"    ✓ Filled {cell_ref}: {str(value)[:50]}")
                    except Exception as e:
                        st.error(f"    ❌ Error filling {cell_ref}: {e}")

    # Handle summary text placement in first sheet
    first_sheet = wb.sheetnames[0]
    ws_first = wb[first_sheet]

    if summary_text:
        st.write(f"📝 Placing summary text in {first_sheet}")
        placed = False
        for row in ws_first.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "skriv her" in cell.value.lower():
                    cell.value = summary_text
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    placed = True
                    st.success(f"  ✓ Summary placed in cell {cell.coordinate}")
                    break
            if placed:
                break

        if not placed:
            ws_first["A46"] = summary_text
            ws_first["A46"].alignment = Alignment(wrap_text=True, vertical="top")
            st.info(f"  ℹ️ Summary placed in default cell A46")

    # Save and return
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    
    st.success("✅ Excel file filled successfully!")
    
    return out.getvalue()


def run():
    """Streamlit page view for the excel filler module"""
    st.title("📊 Excel Filler Module")
    st.write("This module fills Excel templates with extracted data.")
    st.info("Used by the main page to create filled Excel files.")
